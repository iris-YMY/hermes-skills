#!/usr/bin/env python3
"""
数据质量优化脚本 — 退款标记 + 微信退款分类修正 + 微信分类细分 + 亲属卡对齐
在增量写入新数据后运行，对新数据执行质量优化。

用法:
  python3 scripts/data_quality.py --mode all           # 全量执行四项优化
  python3 scripts/data_quality.py --mode refund         # 仅退款标记
  python3 scripts/data_quality.py --mode refund_fix     # 仅微信退款分类修正(支出→不计收支)
  python3 scripts/data_quality.py --mode classify       # 仅微信分类细分
  python3 scripts/data_quality.py --mode family         # 仅亲属卡→家属消费
  python3 scripts/data_quality.py --mode all --wechat_source /path/to/wechat.xlsx  # 含源文件做亲属卡匹配
"""

import argparse
import json
import os
import subprocess
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from collections import defaultdict

CST = timezone(timedelta(hours=8))

# === Config ===
APP_TOKEN = "TcxxbfP05adgltsZpJEcGKi9nme"
TABLE_ID = "tbln6KDEsF2QXyKB"
BATCH_SIZE = 500

# === WeChat 分类关键词映射 ===
CATEGORY_KEYWORDS = {
    '家属消费': ['omd-kimy', '亲属卡', 'kimy'],
    '餐饮': ['美团', '饿了么', '餐饮', '咖啡', '奶茶', '火锅', '烧烤', '小吃',
             'drunk baker', '霸王茶姬', 'chagee', '点餐', '正烧记', '曾三仙',
             '面', '饭', '餐', '茶', '星巴克', 'starbucks', '麦当劳', '肯德基',
             'kfc', '必胜客', 'pizza', '面包', '烘焙', '甜品'],
    '交通': ['滴滴', '高德', '打车', '地铁', '铁路', '12306', '航空', '机票', '火车票',
             '先乘车后付款', '停车费', '出租车', '公交', 'uber'],
    '穿搭美容': ['loro piana', 'gucci', 'stone island', '穿搭', '美容', '护肤', '化妆',
                '服装', '鞋', '恒隆', 'shopping', '奢侈品', '医美', 'spa'],
    '生活日用': ['盒马', '山姆', '超市', '菜', '日用', '家居', '家电',
                '便利店', '全家', 'lawson', '罗森', '7-11', '711',
                '话费', '充值', 'recharge', '余额', '充电宝', '售货机',
                '运费', '快递', '顺丰', '邮政', '洗衣', '干洗', '加油', '洗车'],
    '休闲玩乐': ['电影', '演出', '演唱会', '游乐', '景区', '门票',
                '场地预定', 'ktv', '密室', '剧本杀'],
    '运动': ['体育', '健身', '游泳', '球场', '体育场', '运动',
             '天健体育', '羽毛球', '网球', '篮球', '足球'],
    '宠物': ['宠物', '猫', '狗', 'pet', '兽医', '诊疗', '普普多多宠物'],
    '医疗保健': ['医院', '药', '体检', '医疗', '诊所', '口腔', '牙科',
                '中医', '西医', '药房', '药店'],
    '爱车': ['爱车', '车衣', '车膜', '保养', '维修', '4s'],
}

# 支付宝亲情卡关键词 → 家属消费
ALIPAY_FAMILY_KEYWORDS = ['亲情卡', 'kimy']


def get_user_token():
    token_file = os.path.expanduser("~/.lark/tokens.json")
    with open(token_file) as f:
        return json.load(f)['access_token']


def api_call(method, url, token, payload=None, timeout=30):
    cmd = ['curl', '-s', '-X', method, url,
           '-H', f'Authorization: Bearer {token}',
           '-H', 'Content-Type: application/json']
    tmp = None
    if payload is not None:
        tmp = f'/tmp/feishu_dq_{int(time.time())}.json'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False)
        cmd.extend(['-d', f'@{tmp}'])
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if tmp and os.path.exists(tmp):
        os.remove(tmp)
    return json.loads(result.stdout)


def fetch_all_records(token, platform_filter=None, category_filter=None):
    """Fetch all records from Bitable, optionally filtered."""
    all_records = []
    page_token = None
    while True:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records?page_size=500"
        if page_token:
            url += f"&page_token={page_token}"
        resp = api_call('GET', url, token)
        if resp.get('code') != 0:
            print(f"  ERROR fetching: {resp.get('msg')}")
            break
        items = resp.get('data', {}).get('items', [])
        for rec in items:
            fields = rec.get('fields', {})
            if platform_filter and str(fields.get('平台', '')) != platform_filter:
                continue
            if category_filter and str(fields.get('分类', '')) not in (category_filter if isinstance(category_filter, list) else [category_filter]):
                continue
            all_records.append(rec)
        if not resp.get('data', {}).get('has_more'):
            break
        page_token = resp['data'].get('page_token')
        time.sleep(0.2)
    return all_records


def batch_update(token, updates, label=""):
    """Batch update records."""
    total = len(updates)
    updated = 0
    for i in range(0, total, BATCH_SIZE):
        batch = updates[i:i + BATCH_SIZE]
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records/batch_update"
        resp = api_call('POST', url, token, {"records": batch})
        if resp.get('code') == 0:
            updated += len(batch)
            print(f"    Batch {i // BATCH_SIZE + 1}: OK ({len(batch)} records)")
        else:
            print(f"    Batch {i // BATCH_SIZE + 1}: ERROR - {resp.get('msg', 'unknown')}")
        time.sleep(0.3)
    print(f"  {label}: updated {updated}/{total} records")
    return updated


# ============================================================
# Step 1: 退款交易标记
# ============================================================
def mark_refunds(token):
    """Mark all refund transactions with 是否退款 checkbox."""
    print("\n[1/3] 退款交易标记")
    
    # Check if field exists
    fields_resp = api_call('GET',
        f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/fields", token)
    field_ids = {f['field_name']: f['field_id'] for f in fields_resp.get('data', {}).get('items', [])}
    
    if '是否退款' not in field_ids:
        print("  Adding '是否退款' checkbox field...")
        resp = api_call('POST',
            f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/fields",
            token, {"field_name": "是否退款", "type": 7})
        if resp.get('code') != 0:
            print(f"  ERROR creating field: {resp.get('msg')}")
            return
    
    # Fetch all records
    print("  Fetching all records...")
    all_records = fetch_all_records(token)
    print(f"  Total records: {len(all_records)}")
    
    # Find refund records (skip already marked)
    updates = []
    refund_amount = 0
    for rec in all_records:
        fields = rec.get('fields', {})
        # Skip if already marked
        if fields.get('是否退款'):
            continue
        
        source = str(fields.get('来源', ''))
        category = str(fields.get('分类', ''))
        note = str(fields.get('备注', ''))
        
        is_refund = False
        if '退款' in source or '已全额退款' in source:
            is_refund = True
        if category == '退款':
            is_refund = True
        if '退款' in note:
            is_refund = True
        
        if is_refund:
            updates.append({
                'record_id': rec['record_id'],
                'fields': {'是否退款': True}
            })
            try:
                refund_amount += float(str(fields.get('金额', 0)).replace(',', ''))
            except:
                pass
    
    print(f"  Refund records to mark: {len(updates)}, amount: ¥{refund_amount:,.2f}")
    if updates:
        batch_update(token, updates, "退款标记")
    else:
        print("  No new refund records to mark.")


# ============================================================
# Step 2: 微信「日常消费」细分
# ============================================================
def classify_wechat_daily(token):
    """Re-classify WeChat '日常消费' records into finer categories."""
    print("\n[2/3] 微信「日常消费」细分")
    
    print("  Fetching WeChat '日常消费' records...")
    records = fetch_all_records(token, platform_filter='微信', category_filter='日常消费')
    print(f"  Records to classify: {len(records)}")
    
    updates = []
    cat_count = defaultdict(int)
    
    for rec in records:
        fields = rec.get('fields', {})
        note = str(fields.get('备注', '')).lower()
        
        new_cat = '日常消费'  # default fallback
        for cat, keywords in CATEGORY_KEYWORDS.items():
            if cat == '家属消费':
                continue  # handled in step 3
            if any(kw in note for kw in keywords):
                new_cat = cat
                break
        
        cat_count[new_cat] += 1
        
        if new_cat != '日常消费':
            updates.append({
                'record_id': rec['record_id'],
                'fields': {'分类': new_cat}
            })
    
    print(f"  Classification results:")
    for cat, count in sorted(cat_count.items(), key=lambda x: x[1], reverse=True):
        print(f"    {cat:12s}: {count:4d}")
    print(f"  Records to update: {len(updates)}")
    
    if updates:
        batch_update(token, updates, "分类细分")
    else:
        print("  No records need re-classification.")


# ============================================================
# Step 3: 亲属卡/亲情卡 → 家属消费
# ============================================================
def reclassify_family(token, wechat_source=None):
    """Move 亲属卡/亲情卡/Kimy records to '家属消费' category."""
    print("\n[3/3] 亲属卡/亲情卡 → 家属消费")
    
    updated_total = 0
    
    # 3a: Alipay — match by 备注 field
    print("  [Alipay] Checking '人情社交' records for 亲情卡/Kimy...")
    alipay_social = fetch_all_records(token, platform_filter='支付宝', category_filter=['人情社交', '日常消费'])
    
    alipay_updates = []
    for rec in alipay_social:
        fields = rec.get('fields', {})
        note = str(fields.get('备注', '')).lower()
        if any(kw in note for kw in ALIPAY_FAMILY_KEYWORDS):
            current_cat = str(fields.get('分类', ''))
            if current_cat != '家属消费':
                alipay_updates.append({
                    'record_id': rec['record_id'],
                    'fields': {'分类': '家属消费'}
                })
    
    print(f"    Found {len(alipay_updates)} Alipay records to update")
    if alipay_updates:
        batch_update(token, alipay_updates, "支付宝家属消费")
        updated_total += len(alipay_updates)
    
    # 3b: WeChat — match by timestamp from source file
    if wechat_source and os.path.exists(wechat_source):
        print(f"  [WeChat] Parsing source file: {wechat_source}")
        family_timestamps = extract_family_timestamps(wechat_source)
        print(f"    Family timestamps: {len(family_timestamps)}")
        
        if family_timestamps:
            print("    Fetching WeChat records for timestamp matching...")
            wechat_records = fetch_all_records(token, platform_filter='微信')
            
            wechat_updates = []
            for rec in wechat_records:
                fields = rec.get('fields', {})
                current_cat = str(fields.get('分类', ''))
                if current_cat == '家属消费':
                    continue
                
                ts = fields.get('交易时间')
                try:
                    ts_int = int(ts)
                except:
                    continue
                
                if ts_int in family_timestamps:
                    wechat_updates.append({
                        'record_id': rec['record_id'],
                        'fields': {'分类': '家属消费'}
                    })
            
            print(f"    Matched {len(wechat_updates)} WeChat records")
            if wechat_updates:
                batch_update(token, wechat_updates, "微信家属消费")
                updated_total += len(wechat_updates)
    else:
        # Fallback: match by note field
        print("  [WeChat] No source file, checking 备注 for OMD-Kimy/Kimy...")
        wechat_records = fetch_all_records(token, platform_filter='微信', category_filter=['人情社交', '日常消费'])
        
        wechat_updates = []
        for rec in wechat_records:
            fields = rec.get('fields', {})
            note = str(fields.get('备注', '')).lower()
            source = str(fields.get('来源', '')).lower()
            text = f"{note} {source}"
            if any(kw in text for kw in ['omd-kimy', '亲属卡', 'kimy']):
                current_cat = str(fields.get('分类', ''))
                if current_cat != '家属消费':
                    wechat_updates.append({
                        'record_id': rec['record_id'],
                        'fields': {'分类': '家属消费'}
                    })
        
        print(f"    Found {len(wechat_updates)} WeChat records to update")
        if wechat_updates:
            batch_update(token, wechat_updates, "微信家属消费")
            updated_total += len(wechat_updates)
    
    print(f"\n  Total family records reclassified: {updated_total}")
    return updated_total


def extract_family_timestamps(filepath):
    """Extract timestamps of 亲属卡/Kimy transactions from WeChat source Excel."""
    timestamps = set()
    
    try:
        # Try openpyxl first
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=18, values_only=True):
            if row[0] and isinstance(row[0], datetime):
                trans_type = str(row[1]) if len(row) > 1 and row[1] else ''
                counterparty = str(row[2]) if len(row) > 2 and row[2] else ''
                if '亲属卡' in trans_type or 'OMD-Kimy' in counterparty or 'Kimy' in counterparty:
                    ts = int(row[0].replace(tzinfo=CST).timestamp() * 1000)
                    timestamps.add(ts)
    except (TypeError, Exception):
        # Fallback: parse xlsx as zip
        print("    openpyxl failed, using zipfile fallback...")
        with zipfile.ZipFile(filepath, 'r') as z:
            strings = []
            ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
            if 'xl/sharedStrings.xml' in z.namelist():
                stree = ET.parse(z.open('xl/sharedStrings.xml'))
                for si in stree.findall(f'.//{{{ns}}}si'):
                    texts = si.findall(f'.//{{{ns}}}t')
                    strings.append(''.join(t.text or '' for t in texts))
            
            # Parse sheet1 (WeChat data starts at row 18)
            if 'xl/worksheets/sheet1.xml' in z.namelist():
                stree = ET.parse(z.open('xl/worksheets/sheet1.xml'))
                rows = stree.findall(f'.//{{{ns}}}row')
                for row in rows[17:]:  # skip first 17 rows
                    cells = {}
                    for c in row.findall(f'{{{ns}}}c'):
                        v = c.find(f'{{{ns}}}v')
                        t = c.get('t', '')
                        ref = c.get('r', '')
                        col = ''.join(ch for ch in ref if ch.isalpha())
                        if v is not None and v.text:
                            if t == 's':
                                idx = int(v.text)
                                cells[col] = strings[idx] if idx < len(strings) else ''
                            else:
                                cells[col] = v.text
                    
                    # A=timestamp, B=trans_type, C=counterparty
                    if cells.get('A') and ('亲属卡' in cells.get('B', '') or
                                           'Kimy' in cells.get('C', '') or
                                           'OMD-Kimy' in cells.get('C', '')):
                        try:
                            ts = int(float(cells['A'])) * 1000
                            timestamps.add(ts)
                        except:
                            pass
    
    return timestamps


def main():
    parser = argparse.ArgumentParser(description='数据质量优化')
    parser.add_argument('--mode', choices=['all', 'refund', 'classify', 'family'],
                        default='all', help='执行模式')
    parser.add_argument('--wechat_source', default=None,
                        help='微信原始Excel路径（用于亲属卡时间戳匹配）')
    args = parser.parse_args()

    print("=" * 50)
    print("数据质量优化")
    print("=" * 50)

    token = get_user_token()

    if args.mode in ('all', 'refund'):
        mark_refunds(token)

    if args.mode in ('all', 'classify'):
        classify_wechat_daily(token)

    if args.mode in ('all', 'family'):
        reclassify_family(token, args.wechat_source)

    print("\n✅ 数据质量优化完成!")


if __name__ == '__main__':
    main()
