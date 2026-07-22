#!/usr/bin/env python3
"""
月度财务BI看板自动更新脚本
用法: python3 update_bitable.py --alipay <path> --wechat <path> [--mode full|incremental]
"""

import argparse
import csv
import json
import os
import subprocess
import sys
import time
from datetime import datetime, timezone, timedelta
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# === Config ===
APP_TOKEN = "TcxxbfP05adgltsZpJEcGKi9nme"
TABLE_ID = "tbln6KDEsF2QXyKB"
SUMMARY_TABLE = "tblHqZiC0ZoW1K7o"
SHEET_TOKEN = "V9s7sWj8JhFpRwtIon0cm79Anoc"
SHEET_ID = "28d56d"
BATCH_SIZE = 500
CST = timezone(timedelta(hours=8))

# WeChat category mapping
WX_CATEGORY_MAP = {
    '商户消费': '日常消费', '二维码': '日常消费',
    '亲属卡': '人情社交',
    '红包': '人情社交', '企业微信红包': '人情社交',
    '转账': '转账',
    '退款': '退款',
    '零钱': '投资理财', '提现': '投资理财',
}

def get_user_token():
    token_file = os.path.expanduser("~/.lark/tokens.json")
    with open(token_file) as f:
        return json.load(f)['access_token']

def get_tenant_token():
    app_id = "cli_aa9ebcbfc6e35cba"
    secret_file = os.path.expanduser("~/.lark/app_secret")
    with open(secret_file) as f:
        secret = f.read().strip()
    result = subprocess.run(
        ['curl', '-s', '-X', 'POST',
         'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal',
         '-H', 'Content-Type: application/json',
         '-d', json.dumps({"app_id": app_id, "app_secret": secret})],
        capture_output=True, text=True, timeout=15
    )
    return json.loads(result.stdout)['tenant_access_token']

def api_call(method, url, token, payload=None, timeout=30):
    cmd = ['curl', '-s', '-X', method, url,
           '-H', f'Authorization: Bearer {token}',
           '-H', 'Content-Type: application/json']
    tmp = None
    if payload:
        tmp = f'/tmp/feishu_api_{int(time.time())}.json'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False)
        cmd.extend(['-d', f'@{tmp}'])
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if tmp and os.path.exists(tmp):
        os.remove(tmp)
    return json.loads(result.stdout)

def map_wechat_category(trans_type):
    for key, cat in WX_CATEGORY_MAP.items():
        if key in trans_type:
            return cat
    return '其他'

def map_wechat_direction(direction):
    if direction == '支出': return '支出'
    if direction == '收入': return '收入'
    return '不计收支'

def parse_alipay(filepath):
    records = []
    with open(filepath, 'r', encoding='gbk') as f:
        reader = csv.reader(f)
        rows = list(reader)
    # Find header row to determine column layout
    header_idx = None
    new_format = False
    for i, row in enumerate(rows):
        if row and '交易时间' in row[0]:
            header_idx = i
            cols = row
            # New format (2025+) has 13 cols: 交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注
            # Old format has 9 cols: 记录时间,分类,收支类型,金额,备注,账户,来源,标签
            new_format = len(cols) >= 12 and '交易对方' in cols
            break
    data_start = (header_idx + 1) if header_idx else 11
    for row in rows[data_start:]:
        if len(row) < 7 or not row[0]:
            continue
        try:
            dt = datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S').replace(tzinfo=CST)
            ts = int(dt.timestamp() * 1000)
        except:
            continue
        if new_format:
            # New format: [0]时间 [1]分类 [2]交易对方 [3]对方账号 [4]商品说明 [5]收/支 [6]金额 [7]收/付款方式
            try:
                amt = float(row[6]) if row[6] else 0
            except:
                amt = 0
            records.append({
                "fields": {
                    "交易时间": ts,
                    "分类": row[1],
                    "收支类型": row[5],
                    "金额": amt,
                    "备注": row[4][:200] if len(row) > 4 else '',
                    "账户": row[7] if len(row) > 7 else '',
                    "来源": row[2] if len(row) > 2 else '',
                    "平台": "支付宝"
                }
            })
        else:
            # Old format: [0]时间 [1]分类 [2]收支类型 [3]金额 [4]备注 [5]账户 [6]来源
            try:
                amt = float(row[3]) if row[3] else 0
            except:
                amt = 0
            records.append({
                "fields": {
                    "交易时间": ts,
                    "分类": row[1],
                    "收支类型": row[2],
                    "金额": amt,
                    "备注": row[4][:200] if len(row) > 4 else '',
                    "账户": row[5] if len(row) > 5 else '',
                    "来源": row[6] if len(row) > 6 else '',
                    "平台": "支付宝"
                }
            })
    return records

def parse_wechat(filepath):
    records = []
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=18, values_only=True):
        if not row[0] or not isinstance(row[0], datetime):
            continue
        direction = row[4] if len(row) > 4 else ''
        dtype = map_wechat_direction(direction)
        amt_str = str(row[5]).replace('¥', '').replace('¥', '').strip() if len(row) > 5 and row[5] else '0'
        try:
            amt = float(amt_str)
        except:
            amt = 0
        trans_type = str(row[1]) if len(row) > 1 and row[1] else ''
        cat = map_wechat_category(trans_type)
        note = str(row[3])[:200] if len(row) > 3 and row[3] else ''
        acct = str(row[6]) if len(row) > 6 and row[6] else ''
        src = str(row[7]) if len(row) > 7 and row[7] else ''
        ts = int(row[0].replace(tzinfo=CST).timestamp() * 1000)
        records.append({
            "fields": {
                "交易时间": ts,
                "分类": cat,
                "收支类型": dtype,
                "金额": amt,
                "备注": note,
                "账户": acct,
                "来源": src,
                "平台": "微信"
            }
        })
    return records

def deduplicate(new_records, token):
    """Remove records that already exist in Bitable by matching (timestamp, platform, amount).
    Uses /records endpoint with full pagination. Compares counts per key to handle
    source files that legitimately have multiple transactions at the same second."""
    from collections import Counter
    existing_keys = Counter()
    page_token = None
    while True:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records?page_size=500"
        if page_token:
            url += f"&page_token={page_token}"
        resp = api_call('GET', url, token)
        if resp.get('code') != 0:
            print(f"  Warning: Could not fetch existing records: {resp.get('msg')}")
            return new_records
        for item in resp.get('data', {}).get('items', []):
            fields = item.get('fields', {})
            ts = fields.get('交易时间', 0)
            platform = str(fields.get('平台', ''))
            amt = fields.get('金额', 0)
            try:
                amt = float(amt)
            except:
                amt = 0
            existing_keys[(ts, platform, amt)] += 1
        if not resp.get('data', {}).get('has_more'):
            break
        page_token = resp['data'].get('page_token')
    
    # Count new records by key, only add extras beyond existing count
    new_keys = Counter()
    for r in new_records:
        f = r['fields']
        key = (f['交易时间'], f['平台'], f['金额'])
        new_keys[key] += 1
    
    filtered = []
    seen = Counter()
    for r in new_records:
        f = r['fields']
        key = (f['交易时间'], f['平台'], f['金额'])
        existing_count = existing_keys.get(key, 0)
        seen_count = seen.get(key, 0)
        # Allow up to (source_count - existing_count) copies through
        # source_count = new_keys[key], so extras = new_keys[key] - (existing_count already in bitable)
        if seen_count < max(new_keys[key] - existing_count, 0):
            filtered.append(r)
        seen[key] = seen_count + 1
    
    skipped = len(new_records) - len(filtered)
    print(f"  Dedup: {len(new_records)} → {len(filtered)} new (skipped {skipped} existing)")
    return filtered

def batch_insert(records, token):
    """Insert records in batches of 500"""
    total = len(records)
    inserted = 0
    for i in range(0, total, BATCH_SIZE):
        batch = records[i:i+BATCH_SIZE]
        payload = {"records": batch}
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records/batch_create"
        resp = api_call('POST', url, token, payload)
        if resp.get('code') == 0:
            inserted += len(batch)
            print(f"  Batch {i//BATCH_SIZE + 1}: OK ({len(batch)} records)")
        else:
            print(f"  Batch {i//BATCH_SIZE + 1}: ERROR - {resp.get('msg', 'unknown')}")
        time.sleep(0.3)
    return inserted

def update_summary(token):
    """Recalculate and update monthly summary table"""
    # Fetch all records
    all_records = []
    page_token = None
    while True:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records?page_size=500"
        if page_token:
            url += f"&page_token={page_token}"
        resp = api_call('GET', url, token)
        if resp.get('code') != 0:
            print(f"  Error fetching records for summary: {resp.get('msg')}")
            return
        all_records.extend(resp.get('data', {}).get('items', []))
        if not resp.get('data', {}).get('has_more'):
            break
        page_token = resp['data'].get('page_token')
    
    # Aggregate by month
    monthly = defaultdict(lambda: {'收入': 0, '支出': 0})
    for rec in all_records:
        fields = rec.get('fields', {})
        ts = fields.get('交易时间', 0)
        if not ts:
            continue
        dt = datetime.fromtimestamp(ts / 1000, tz=CST)
        month_key = dt.strftime('%Y-%m')
        dtype = fields.get('收支类型', '')
        amt = fields.get('金额', 0) or 0
        try:
            amt = float(amt)
        except (ValueError, TypeError):
            amt = 0
        if dtype == '收入':
            monthly[month_key]['收入'] += amt
        elif dtype == '支出':
            monthly[month_key]['支出'] += amt
    
    # Build summary records (only monthly rows, no 年度合计/月均)
    summary_records = []
    for m in sorted(monthly.keys()):
        inc = round(monthly[m]['收入'], 2)
        exp = round(monthly[m]['支出'], 2)
        net = round(inc - exp, 2)
        sr = f"{(net/inc*100):.1f}%" if inc > 0 else "N/A"
        summary_records.append({"fields": {"月份": m, "收入": inc, "支出": exp, "净额": net, "储蓄率": sr}})
    
    # Delete old summary records and insert new ones
    # Get existing record IDs
    existing_ids = []
    page_token = None
    while True:
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{SUMMARY_TABLE}/records?page_size=500"
        if page_token:
            url += f"&page_token={page_token}"
        resp = api_call('GET', url, token)
        if resp.get('code') != 0:
            break
        for item in resp.get('data', {}).get('items', []):
            existing_ids.append(item['record_id'])
        if not resp.get('data', {}).get('has_more'):
            break
        page_token = resp['data'].get('page_token')
    
    # Batch delete old records
    if existing_ids:
        for i in range(0, len(existing_ids), BATCH_SIZE):
            batch_ids = existing_ids[i:i+BATCH_SIZE]
            api_call('POST', f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{SUMMARY_TABLE}/records/batch_delete",
                     token, {"records": batch_ids})
            time.sleep(0.2)
    
    # Insert new summary
    for i in range(0, len(summary_records), BATCH_SIZE):
        batch = summary_records[i:i+BATCH_SIZE]
        api_call('POST', f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{SUMMARY_TABLE}/records/batch_create",
                 token, {"records": batch})
        time.sleep(0.2)
    
    print(f"  Summary updated: {len(summary_records)} rows")

def get_sheet_row_count(token):
    """Get current row count of the spreadsheet"""
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{SHEET_TOKEN}/metainfo"
    resp = api_call('GET', url, token)
    if resp.get('code') != 0:
        print(f"  Warning: Could not get sheet info: {resp.get('msg')}")
        return 4223  # fallback to known count
    sheets = resp.get('data', {}).get('sheets', [])
    for sheet in sheets:
        if sheet.get('sheetId') == SHEET_ID:
            return sheet.get('properties', {}).get('rowCount', 4223)
    return 4223

def sync_to_spreadsheet(records, token):
    """Append new records to spreadsheet"""
    if not records:
        print("  No records to sync")
        return
    
    # Get current row count
    current_rows = get_sheet_row_count(token)
    print(f"  Current sheet rows: {current_rows}")
    
    # Convert records to sheet format
    sheet_data = []
    for rec in records:
        fields = rec.get('fields', {})
        ts = fields.get('交易时间', 0)
        if ts:
            dt = datetime.fromtimestamp(ts / 1000, tz=CST)
            time_str = dt.strftime('%Y-%m-%d %H:%M:%S')
        else:
            time_str = ''
        
        sheet_data.append([
            time_str,
            fields.get('分类', ''),
            fields.get('收支类型', ''),
            fields.get('金额', 0),
            fields.get('备注', ''),
            fields.get('账户', ''),
            fields.get('来源', ''),
            fields.get('平台', '')
        ])
    
    # Append in batches
    total = len(sheet_data)
    appended = 0
    for i in range(0, total, BATCH_SIZE):
        batch = sheet_data[i:i+BATCH_SIZE]
        start_row = current_rows + i + 1
        end_row = start_row + len(batch) - 1
        range_str = f"{SHEET_ID}!A{start_row}:H{end_row}"
        
        payload = {
            "valueRange": {
                "range": range_str,
                "values": batch
            }
        }
        
        url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{SHEET_TOKEN}/values"
        resp = api_call('PUT', url, token, payload)
        if resp.get('code') == 0:
            appended += len(batch)
            print(f"  Sheet batch {i//BATCH_SIZE + 1}: OK ({len(batch)} rows, A{start_row}:H{end_row})")
        else:
            print(f"  Sheet batch {i//BATCH_SIZE + 1}: ERROR - {resp.get('msg', 'unknown')}")
        time.sleep(0.3)
    
    print(f"  Sheet synced: {appended}/{total} rows")

def send_notification(token, stats):
    """Send update notification card"""
    tenant_token = get_tenant_token()
    card = {
        "config": {"wide_screen_mode": True},
        "header": {"title": {"tag": "plain_text", "content": "✅ 财务看板数据已更新"}, "template": "green"},
        "elements": [
            {"tag": "div", "text": {"tag": "lark_md", "content": f"**更新摘要**\n\n• 新增记录：{stats['new_records']} 条\n• 支付宝：{stats['alipay_count']} 条\n• 微信：{stats['wechat_count']} 条\n• 跳过重复：{stats['skipped']} 条\n• 汇总表已刷新\n• 总记录数：{stats['total_records']} 条"}},
            {"tag": "hr"},
            {"tag": "note", "elements": [{"tag": "plain_text", "content": f"更新时间：{datetime.now(CST).strftime('%Y-%m-%d %H:%M')}"}]},
            {"tag": "action", "actions": [{"tag": "button", "text": {"tag": "plain_text", "content": "📊 打开BI看板"}, "url": f"https://e1kg6bc4dl9.feishu.cn/base/{APP_TOKEN}", "type": "primary"}]}
        ]
    }
    api_call('POST', 'https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id',
             tenant_token, {"receive_id": "oc_d811c650f76f16e98ac7a65517e0128f", "msg_type": "interactive", "content": json.dumps(card, ensure_ascii=False)})

def main():
    parser = argparse.ArgumentParser(description='月度财务看板更新')
    parser.add_argument('--alipay', required=True, help='支付宝CSV路径')
    parser.add_argument('--wechat', required=True, help='微信Excel路径')
    parser.add_argument('--mode', choices=['full', 'incremental'], default='incremental', help='更新模式')
    args = parser.parse_args()

    print("=" * 50)
    print("月度财务BI看板自动更新")
    print("=" * 50)

    # Step 1: Parse files
    print("\n[1/4] 解析账单文件...")
    alipay_records = parse_alipay(args.alipay)
    wechat_records = parse_wechat(args.wechat)
    print(f"  支付宝: {len(alipay_records)} 条")
    print(f"  微信: {len(wechat_records)} 条")
    
    all_new = alipay_records + wechat_records
    print(f"  合计: {len(all_new)} 条")

    # Step 2: Get token & dedup
    print("\n[2/4] 去重检查...")
    token = get_user_token()
    if args.mode == 'incremental':
        filtered = deduplicate(all_new, token)
    else:
        filtered = all_new
        print("  Full mode: no dedup")

    skipped = len(all_new) - len(filtered)

    # Step 3: Insert to Bitable
    print("\n[3/5] 写入多维表格...")
    if filtered:
        inserted = batch_insert(filtered, token)
        print(f"  成功写入: {inserted} 条")
    else:
        inserted = 0
        print("  无新记录需要写入")

    # Step 4: Sync to spreadsheet
    print("\n[4/5] 同步电子表格...")
    if filtered:
        sync_to_spreadsheet(filtered, token)
    else:
        print("  无新记录需要同步")

    # Step 5: Update summary
    print("\n[5/5] 更新月度汇总...")
    update_summary(token)

    # Send notification
    print("\n发送通知...")
    stats = {
        'new_records': inserted,
        'alipay_count': len(alipay_records),
        'wechat_count': len(wechat_records),
        'skipped': skipped,
        'total_records': 'checking...'
    }
    send_notification(token, stats)
    print("\n✅ 更新完成!")

if __name__ == '__main__':
    main()
